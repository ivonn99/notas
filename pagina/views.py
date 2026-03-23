from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import csrf_protect
from django.http import JsonResponse
from datetime import datetime
import logging
import time
from .decorators import rol_requerido

logger = logging.getLogger(__name__)


@csrf_protect
def login_view(request):
    """Vista de login"""
    from .models import Usuario
    
    if request.user.is_authenticated:
        return redirect('pagina_principal')
    
    context = {}
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        # Siempre mantener el username en el contexto para que no se limpie
        context['username_value'] = username
        
        # Verificar si el usuario existe
        try:
            usuario_existe = Usuario.objects.filter(username=username).exists()
        except:
            usuario_existe = False
        
        if not usuario_existe:
            # Usuario no existe: mantener el username, no limpiar nada
            messages.error(request, 'El usuario no existe. Verifica que hayas ingresado correctamente tu nombre de usuario.')
        else:
            # El usuario existe, verificar contraseña
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                if user.activo:
                    login(request, user)
                    messages.success(request, f'Bienvenido, {user.nombre_completo}, {user.get_rol_display()}')
                    
                    # Redirigir según el rol
                    next_url = request.GET.get('next', 'pagina_principal')
                    return redirect(next_url)
                else:
                    messages.error(request, 'Tu cuenta está desactivada. Contacta al administrador.')
            else:
                # Contraseña incorrecta: mantener el username, limpiar solo la contraseña
                messages.error(request, 'La contraseña es incorrecta. Verifica que hayas ingresado correctamente tu contraseña.')
                # No agregar password_value al contexto para que se limpie
    
    # Asegurar que el contexto tenga el request para CSRF
    return render(request, 'pagina/login.html', context)


def logout_view(request):
    """Vista de logout"""
    logout(request)
    messages.success(request, 'Has cerrado sesión correctamente.')
    return redirect('login')


def healthz(request):
    """Health check básico para monitoreo."""
    return JsonResponse({'status': 'ok'})


@login_required
def pagina_principal(request):
    """Página Principal"""
    context = {
        'titulo': 'Página Principal',
        'contenido': 'Panel principal del sistema de gestión de notas de crédito.'
    }
    return render(request, 'pagina/pagina.html', context)


@login_required
def todas_las_notas(request):
    """Todas las Notas de Crédito con Scroll Infinito"""
    from .models import NotaCredito, Ruta
    from django.db.models import Q, Sum, Count, F, ExpressionWrapper, IntegerField
    from django.db.models.functions import Extract
    from decimal import Decimal
    from django.http import JsonResponse
    import json
    
    # Si es una petición AJAX para cargar más notas
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('ajax') == '1':
        page = int(request.GET.get('page', 1))
        per_page = 50
        offset = (page - 1) * per_page
        
        # Obtener todas las notas con filtros
        notas_query = NotaCredito.objects.all().select_related('ruta', 'usuario')
        
        # Filtros
        estado_filtro = request.GET.get('estado', '')
        empresa_filtro = request.GET.get('empresa', '')
        ruta_filtro = request.GET.get('ruta', '')
        busqueda = request.GET.get('busqueda', '').strip()
        dias_filtro = request.GET.get('dias', '').strip()
        dias_operador = request.GET.get('dias_operador', 'mayor')
        ordenar_por = request.GET.get('ordenar_por', 'fecha_nota_desc')  # 'fecha_nota_desc', 'fecha_nota_asc', 'diferencia_dias_desc', 'diferencia_dias_asc'
        
        # Aplicar filtros
        if estado_filtro:
            notas_query = notas_query.filter(estado=estado_filtro)
        
        if empresa_filtro:
            notas_query = notas_query.filter(empresa=empresa_filtro)
        
        if ruta_filtro:
            notas_query = notas_query.filter(ruta_id=ruta_filtro)
        
        if busqueda:
            notas_query = notas_query.filter(
                Q(serie_folio__icontains=busqueda) |
                Q(cliente__icontains=busqueda) |
                Q(usuario_vendedor_pv__icontains=busqueda)
            )
        
        # Filtro por diferencia de días (valor absoluto)
        if dias_filtro:
            try:
                dias_valor = int(dias_filtro)
                # En PostgreSQL, la resta de DATE devuelve INTEGER directamente
                # Usamos SQL raw para convertir a días correctamente
                from django.db.models import Func
                # Usar EXTRACT(EPOCH FROM interval) / 86400 para convertir a días
                diferencia = Func(
                    F('fecha_corriente') - F('fecha_nota'),
                    function='EXTRACT',
                    template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                    output_field=IntegerField()
                )
                diferencia_inversa = Func(
                    F('fecha_nota') - F('fecha_corriente'),
                    function='EXTRACT',
                    template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                    output_field=IntegerField()
                )
                if dias_operador == 'mayor':
                    # Mayor a X días: cualquiera de las dos diferencias debe ser mayor a X
                    notas_query = notas_query.annotate(
                        diferencia_dias=diferencia,
                        diferencia_inversa=diferencia_inversa
                    ).filter(
                        Q(diferencia_dias__gt=dias_valor) | Q(diferencia_inversa__gt=dias_valor)
                    )
                else:  # menor
                    # Menor a X días: ambas diferencias deben estar entre -X y X
                    notas_query = notas_query.annotate(
                        diferencia_dias=diferencia,
                        diferencia_inversa=diferencia_inversa
                    ).filter(
                        (Q(diferencia_dias__lt=dias_valor) & Q(diferencia_dias__gt=-dias_valor)) |
                        (Q(diferencia_inversa__lt=dias_valor) & Q(diferencia_inversa__gt=-dias_valor))
                    )
            except ValueError:
                pass  # Si no es un número válido, ignorar el filtro
        
        # Si es VENDEDOR, solo mostrar sus notas o las de sus rutas
        if request.user.rol == 'VENDEDOR':
            rutas_usuario = request.user.rutas.all()
            if rutas_usuario:
                notas_query = notas_query.filter(ruta__in=rutas_usuario)
            else:
                notas_query = notas_query.none()
        
        # Aplicar ordenamiento
        from django.db.models import Func
        if ordenar_por == 'diferencia_dias_desc':
            # Ordenar por diferencia de días (valor absoluto) descendente
            notas_query = notas_query.annotate(
                diferencia_dias_abs=Func(
                    Func(
                        F('fecha_corriente') - F('fecha_nota'),
                        function='EXTRACT',
                        template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                        output_field=IntegerField()
                    ),
                    function='ABS',
                    template="ABS(%(expressions)s)",
                    output_field=IntegerField()
                )
            ).order_by('-diferencia_dias_abs', '-fecha_nota', '-created_at')
        elif ordenar_por == 'diferencia_dias_asc':
            # Ordenar por diferencia de días (valor absoluto) ascendente
            notas_query = notas_query.annotate(
                diferencia_dias_abs=Func(
                    Func(
                        F('fecha_corriente') - F('fecha_nota'),
                        function='EXTRACT',
                        template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                        output_field=IntegerField()
                    ),
                    function='ABS',
                    template="ABS(%(expressions)s)",
                    output_field=IntegerField()
                )
            ).order_by('diferencia_dias_abs', 'fecha_nota', 'created_at')
        elif ordenar_por == 'fecha_nota_asc':
            # Ordenar por fecha nota ascendente
            notas_query = notas_query.order_by('fecha_nota', 'created_at')
        else:  # ordenar_por == 'fecha_nota_desc' (por defecto)
            notas_query = notas_query.order_by('-fecha_nota', '-created_at')
        
        # Obtener el total antes de paginar
        total_count = notas_query.count()
        
        # Paginar
        notas = notas_query[offset:offset + per_page]
        
        # Serializar notas
        notas_data = []
        for nota in notas:
            notas_data.append({
                'id': nota.id,
                'serie_folio': nota.serie_folio,
                'empresa': nota.get_empresa_display(),
                'empresa_value': nota.empresa,
                'fecha_nota': nota.fecha_nota.strftime('%d/%m/%Y'),
                'cliente': nota.cliente,
                'ruta_codigo': nota.ruta.codigo,
                'ruta_nombre': nota.ruta.nombre,
                'usuario_vendedor_pv': nota.usuario_vendedor_pv or '',
                'monto': str(nota.monto),
                'abono': str(nota.abono),
                'saldo': str(nota.saldo),
                'estado': nota.get_estado_display(),
                'estado_value': nota.estado,
                'fecha_corriente': nota.fecha_corriente.strftime('%d/%m/%Y'),
                'fecha_corriente_iso': nota.fecha_corriente.isoformat(),
                'diferencia_dias': abs((nota.fecha_corriente - nota.fecha_nota).days),
                'requiere_atencion': nota.requiere_atencion,
            })
        
        has_more = (offset + per_page) < total_count
        
        return JsonResponse({
            'notas': notas_data,
            'has_more': has_more,
            'page': page,
            'total': total_count
        })
    
    # Vista normal (primera carga)
    # Obtener todas las notas con cálculo de diferencia de días
    notas_query = NotaCredito.objects.all().select_related('ruta', 'usuario')
    
    # Filtros
    estado_filtro = request.GET.get('estado', '')
    empresa_filtro = request.GET.get('empresa', '')
    ruta_filtro = request.GET.get('ruta', '')
    busqueda = request.GET.get('busqueda', '').strip()
    dias_filtro = request.GET.get('dias', '').strip()
    dias_operador = request.GET.get('dias_operador', 'mayor')  # 'mayor' o 'menor'
    ordenar_por = request.GET.get('ordenar_por', 'fecha_nota_desc')  # 'fecha_nota_desc', 'fecha_nota_asc', 'diferencia_dias_desc', 'diferencia_dias_asc'
    
    # Aplicar filtros
    if estado_filtro:
        notas_query = notas_query.filter(estado=estado_filtro)
    
    if empresa_filtro:
        notas_query = notas_query.filter(empresa=empresa_filtro)
    
    if ruta_filtro:
        notas_query = notas_query.filter(ruta_id=ruta_filtro)
    
    if busqueda:
        notas_query = notas_query.filter(
            Q(serie_folio__icontains=busqueda) |
            Q(cliente__icontains=busqueda) |
            Q(usuario_vendedor_pv__icontains=busqueda)
        )
    
    # Filtro por diferencia de días (valor absoluto)
    if dias_filtro:
        try:
            dias_valor = int(dias_filtro)
            # En PostgreSQL, la resta de DATE devuelve INTEGER directamente
            # Pero Django lo trata como intervalo, así que usamos EXTRACT(EPOCH) para convertir a días
            from django.db.models import Func
            # Convertir intervalo a días usando EXTRACT(EPOCH FROM interval) / 86400
            diferencia = Func(
                F('fecha_corriente') - F('fecha_nota'),
                function='EXTRACT',
                template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                output_field=IntegerField()
            )
            diferencia_inversa = Func(
                F('fecha_nota') - F('fecha_corriente'),
                function='EXTRACT',
                template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                output_field=IntegerField()
            )
            if dias_operador == 'mayor':
                # Mayor a X días: cualquiera de las dos diferencias debe ser mayor a X
                notas_query = notas_query.annotate(
                    diferencia_dias=diferencia,
                    diferencia_inversa=diferencia_inversa
                ).filter(
                    Q(diferencia_dias__gt=dias_valor) | Q(diferencia_inversa__gt=dias_valor)
                )
            else:  # menor
                # Menor a X días: ambas diferencias deben estar entre -X y X
                notas_query = notas_query.annotate(
                    diferencia_dias=diferencia,
                    diferencia_inversa=diferencia_inversa
                ).filter(
                    (Q(diferencia_dias__lt=dias_valor) & Q(diferencia_dias__gt=-dias_valor)) |
                    (Q(diferencia_inversa__lt=dias_valor) & Q(diferencia_inversa__gt=-dias_valor))
                )
        except ValueError:
            pass  # Si no es un número válido, ignorar el filtro
    
    # Si es VENDEDOR, solo mostrar sus notas o las de sus rutas
    if request.user.rol == 'VENDEDOR':
        rutas_usuario = request.user.rutas.all()
        if rutas_usuario:
            notas_query = notas_query.filter(ruta__in=rutas_usuario)
        else:
            notas_query = notas_query.none()
    
    # Aplicar ordenamiento
    from django.db.models import Func
    if ordenar_por == 'diferencia_dias_desc':
        # Ordenar por diferencia de días (valor absoluto) descendente
        notas_query = notas_query.annotate(
            diferencia_dias_abs=Func(
                Func(
                    F('fecha_corriente') - F('fecha_nota'),
                    function='EXTRACT',
                    template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                    output_field=IntegerField()
                ),
                function='ABS',
                template="ABS(%(expressions)s)",
                output_field=IntegerField()
            )
        ).order_by('-diferencia_dias_abs', '-fecha_nota', '-created_at')
    elif ordenar_por == 'diferencia_dias_asc':
        # Ordenar por diferencia de días (valor absoluto) ascendente
        notas_query = notas_query.annotate(
            diferencia_dias_abs=Func(
                Func(
                    F('fecha_corriente') - F('fecha_nota'),
                    function='EXTRACT',
                    template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                    output_field=IntegerField()
                ),
                function='ABS',
                template="ABS(%(expressions)s)",
                output_field=IntegerField()
            )
        ).order_by('diferencia_dias_abs', 'fecha_nota', 'created_at')
    elif ordenar_por == 'fecha_nota_asc':
        # Ordenar por fecha nota ascendente
        notas_query = notas_query.order_by('fecha_nota', 'created_at')
    else:  # ordenar_por == 'fecha_nota_desc' (por defecto)
        notas_query = notas_query.order_by('-fecha_nota', '-created_at')
    
    # Obtener solo las primeras 50 para la carga inicial
    notas = notas_query[:50]
    
    # Estadísticas (usando el query completo, no solo las 50)
    total_notas = notas_query.count()
    total_monto = notas_query.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
    total_saldo = notas_query.aggregate(Sum('saldo'))['saldo__sum'] or Decimal('0.00')
    
    # Contar por estado
    pendientes = notas_query.filter(estado='PENDIENTE').count()
    resueltas = notas_query.filter(estado='RESUELTA').count()
    canceladas = notas_query.filter(estado='CANCELADA').count()
    
    # Obtener rutas para el filtro
    rutas = Ruta.objects.filter(activa=True).order_by('codigo')
    
    # Verificar si hay más notas
    has_more = total_notas > 50
    
    context = {
        'notas': notas,
        'total_notas': total_notas,
        'total_monto': total_monto,
        'total_saldo': total_saldo,
        'pendientes': pendientes,
        'resueltas': resueltas,
        'canceladas': canceladas,
        'rutas': rutas,
        'estado_filtro': estado_filtro,
        'empresa_filtro': empresa_filtro,
        'ruta_filtro': ruta_filtro,
        'busqueda': busqueda,
        'dias_filtro': dias_filtro,
        'dias_operador': dias_operador,
        'ordenar_por': ordenar_por,
        'has_more': has_more,
    }
    return render(request, 'pagina/todas_las_notas.html', context)


@rol_requerido('CREDITO', 'ADMIN')
def alertas(request):
    """Alertas de Crédito y Cobranza"""
    context = {
        'titulo': 'Alertas',
        'contenido': 'Notas que reaparecen o requieren atención especial.'
    }
    return render(request, 'pagina/pagina.html', context)


@rol_requerido('CREDITO', 'ADMIN', 'VENDEDOR')
def seguimiento(request):
    """Seguimiento de Notas - Lista de notas para seguimiento"""
    from .models import NotaCredito, Ruta
    from django.db.models import Q, F, ExpressionWrapper, IntegerField
    from django.db.models.functions import Extract
    from django.db.models import Func
    from django.http import JsonResponse
    
    # Si es una petición AJAX para cargar más notas
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('ajax') == '1':
        page = int(request.GET.get('page', 1))
        per_page = 50
        offset = (page - 1) * per_page
        
        # Filtrar notas según el rol del usuario
        notas_query = NotaCredito.objects.all().select_related('ruta', 'usuario')
        
        # Si es VENDEDOR, solo mostrar notas de sus rutas asignadas
        if request.user.rol == 'VENDEDOR':
            rutas_usuario = request.user.rutas.all()
            notas_query = notas_query.filter(ruta__in=rutas_usuario)
        
        # Filtros
        estado_filtro = request.GET.get('estado', '')
        empresa_filtro = request.GET.get('empresa', '')
        ruta_filtro = request.GET.get('ruta', '')
        busqueda = request.GET.get('busqueda', '').strip()
        dias_filtro = request.GET.get('dias', '').strip()
        dias_operador = request.GET.get('dias_operador', 'mayor')
        ordenar_por = request.GET.get('ordenar_por', 'fecha_nota_desc')
        requiere_atencion_filtro = request.GET.get('requiere_atencion', '') == '1'
        
        # Aplicar filtros
        if estado_filtro:
            notas_query = notas_query.filter(estado=estado_filtro)
        
        if empresa_filtro:
            notas_query = notas_query.filter(empresa=empresa_filtro)
        
        if ruta_filtro:
            notas_query = notas_query.filter(ruta_id=ruta_filtro)
        
        if busqueda:
            notas_query = notas_query.filter(
                Q(serie_folio__icontains=busqueda) |
                Q(cliente__icontains=busqueda)
            )
        
        # Filtro por diferencia de días
        if dias_filtro:
            try:
                dias_valor = int(dias_filtro)
                diferencia = Func(
                    F('fecha_corriente') - F('fecha_nota'),
                    function='EXTRACT',
                    template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                    output_field=IntegerField()
                )
                diferencia_inversa = Func(
                    F('fecha_nota') - F('fecha_corriente'),
                    function='EXTRACT',
                    template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                    output_field=IntegerField()
                )
                if dias_operador == 'mayor':
                    notas_query = notas_query.annotate(
                        diferencia_dias=diferencia,
                        diferencia_inversa=diferencia_inversa
                    ).filter(
                        Q(diferencia_dias__gt=dias_valor) | Q(diferencia_inversa__gt=dias_valor)
                    )
                else:
                    notas_query = notas_query.annotate(
                        diferencia_dias=diferencia,
                        diferencia_inversa=diferencia_inversa
                    ).filter(
                        (Q(diferencia_dias__lt=dias_valor) & Q(diferencia_dias__gt=-dias_valor)) |
                        (Q(diferencia_inversa__lt=dias_valor) & Q(diferencia_inversa__gt=-dias_valor))
                    )
            except ValueError:
                pass
        
        # Filtro de notas que requieren atención
        if requiere_atencion_filtro:
            notas_query = notas_query.filter(requiere_atencion=True)
        
        # Aplicar ordenamiento
        if ordenar_por == 'diferencia_dias_desc':
            notas_query = notas_query.annotate(
                diferencia_dias_abs=Func(
                    Func(
                        F('fecha_corriente') - F('fecha_nota'),
                        function='EXTRACT',
                        template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                        output_field=IntegerField()
                    ),
                    function='ABS',
                    template="ABS(%(expressions)s)",
                    output_field=IntegerField()
                )
            ).order_by('-diferencia_dias_abs', '-fecha_nota', '-created_at')
        elif ordenar_por == 'diferencia_dias_asc':
            notas_query = notas_query.annotate(
                diferencia_dias_abs=Func(
                    Func(
                        F('fecha_corriente') - F('fecha_nota'),
                        function='EXTRACT',
                        template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                        output_field=IntegerField()
                    ),
                    function='ABS',
                    template="ABS(%(expressions)s)",
                    output_field=IntegerField()
                )
            ).order_by('diferencia_dias_abs', 'fecha_nota', 'created_at')
        elif ordenar_por == 'fecha_nota_asc':
            notas_query = notas_query.order_by('fecha_nota', 'created_at')
        else:
            notas_query = notas_query.order_by('-fecha_nota', '-created_at')
        
        # Obtener total antes de paginación
        total_count = notas_query.count()
        
        # Aplicar paginación
        notas = notas_query[offset:offset + per_page]
        
        # Preparar datos para JSON
        notas_data = []
        for nota in notas:
            dias_diff = abs((nota.fecha_corriente - nota.fecha_nota).days)
            notas_data.append({
                'id': nota.id,
                'serie_folio': nota.serie_folio,
                'empresa': nota.get_empresa_display(),
                'empresa_value': nota.empresa,
                'fecha_nota': nota.fecha_nota.strftime('%d/%m/%Y'),
                'cliente': nota.cliente,
                'ruta_codigo': nota.ruta.codigo,
                'ruta_nombre': nota.ruta.nombre,
                'usuario_vendedor_pv': nota.usuario_vendedor_pv or '',
                'monto': str(nota.monto),
                'abono': str(nota.abono),
                'saldo': str(nota.saldo),
                'estado': nota.get_estado_display(),
                'estado_value': nota.estado,
                'fecha_corriente': nota.fecha_corriente.strftime('%d/%m/%Y'),
                'diferencia_dias': dias_diff,
                'requiere_atencion': nota.requiere_atencion,
            })
        
        has_more = (offset + per_page) < total_count
        
        return JsonResponse({
            'notas': notas_data,
            'has_more': has_more,
            'page': page,
            'total': total_count
        })
    
    # Código para carga inicial (no AJAX)
    # Filtrar notas según el rol del usuario
    notas_query = NotaCredito.objects.all().select_related('ruta', 'usuario')
    
    # Si es VENDEDOR, solo mostrar notas de sus rutas asignadas
    if request.user.rol == 'VENDEDOR':
        rutas_usuario = request.user.rutas.all()
        notas_query = notas_query.filter(ruta__in=rutas_usuario)
    
    # Filtros
    estado_filtro = request.GET.get('estado', '')
    empresa_filtro = request.GET.get('empresa', '')
    ruta_filtro = request.GET.get('ruta', '')
    busqueda = request.GET.get('busqueda', '').strip()
    dias_filtro = request.GET.get('dias', '').strip()
    dias_operador = request.GET.get('dias_operador', 'mayor')
    ordenar_por = request.GET.get('ordenar_por', 'fecha_nota_desc')
    requiere_atencion_filtro = request.GET.get('requiere_atencion', '') == '1'  # Filtro para notas que requieren atención
    
    if estado_filtro:
        notas_query = notas_query.filter(estado=estado_filtro)
    
    if empresa_filtro:
        notas_query = notas_query.filter(empresa=empresa_filtro)
    
    if ruta_filtro:
        notas_query = notas_query.filter(ruta_id=ruta_filtro)
    
    if busqueda:
        notas_query = notas_query.filter(
            Q(serie_folio__icontains=busqueda) |
            Q(cliente__icontains=busqueda)
        )
    
    # Filtro por diferencia de días (valor absoluto)
    if dias_filtro:
        try:
            dias_valor = int(dias_filtro)
            diferencia = Func(
                F('fecha_corriente') - F('fecha_nota'),
                function='EXTRACT',
                template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                output_field=IntegerField()
            )
            diferencia_inversa = Func(
                F('fecha_nota') - F('fecha_corriente'),
                function='EXTRACT',
                template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                output_field=IntegerField()
            )
            if dias_operador == 'mayor':
                notas_query = notas_query.annotate(
                    diferencia_dias=diferencia,
                    diferencia_inversa=diferencia_inversa
                ).filter(
                    Q(diferencia_dias__gt=dias_valor) | Q(diferencia_inversa__gt=dias_valor)
                )
            else:  # menor
                notas_query = notas_query.annotate(
                    diferencia_dias=diferencia,
                    diferencia_inversa=diferencia_inversa
                ).filter(
                    (Q(diferencia_dias__lt=dias_valor) & Q(diferencia_dias__gt=-dias_valor)) |
                    (Q(diferencia_inversa__lt=dias_valor) & Q(diferencia_inversa__gt=-dias_valor))
                )
        except ValueError:
            pass
    
    # Aplicar ordenamiento
    if ordenar_por == 'diferencia_dias_desc':
        notas_query = notas_query.annotate(
            diferencia_dias_abs=Func(
                Func(
                    F('fecha_corriente') - F('fecha_nota'),
                    function='EXTRACT',
                    template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                    output_field=IntegerField()
                ),
                function='ABS',
                template="ABS(%(expressions)s)",
                output_field=IntegerField()
            )
        ).order_by('-diferencia_dias_abs', '-fecha_nota', '-created_at')
    elif ordenar_por == 'diferencia_dias_asc':
        notas_query = notas_query.annotate(
            diferencia_dias_abs=Func(
                Func(
                    F('fecha_corriente') - F('fecha_nota'),
                    function='EXTRACT',
                    template="EXTRACT(EPOCH FROM (%(expressions)s))::integer / 86400",
                    output_field=IntegerField()
                ),
                function='ABS',
                template="ABS(%(expressions)s)",
                output_field=IntegerField()
            )
        ).order_by('diferencia_dias_abs', 'fecha_nota', 'created_at')
    elif ordenar_por == 'fecha_nota_asc':
        notas_query = notas_query.order_by('fecha_nota', 'created_at')
    else:  # ordenar_por == 'fecha_nota_desc' (por defecto)
        notas_query = notas_query.order_by('-fecha_nota', '-created_at')
    
    # Obtener rutas para el filtro
    if request.user.rol == 'VENDEDOR':
        rutas = request.user.rutas.filter(activa=True).order_by('codigo')
    else:
        rutas = Ruta.objects.filter(activa=True).order_by('codigo')
    
    # Filtro de notas que requieren atención
    if requiere_atencion_filtro:
        notas_query = notas_query.filter(requiere_atencion=True)
    
    # Obtener total antes de paginación
    total_notas = notas_query.count()
    
    # Obtener solo las primeras 50 para la carga inicial
    notas = notas_query[:50]
    
    # Verificar si hay más notas
    has_more = total_notas > 50
    
    context = {
        'titulo': 'Seguimiento',
        'notas': notas,
        'total_notas': total_notas,
        'has_more': has_more,
        'rutas': rutas,
        'estado_filtro': estado_filtro,
        'empresa_filtro': empresa_filtro,
        'ruta_filtro': ruta_filtro,
        'busqueda': busqueda,
        'dias_filtro': dias_filtro,
        'dias_operador': dias_operador,
        'ordenar_por': ordenar_por,
        'requiere_atencion_filtro': requiere_atencion_filtro,
    }
    return render(request, 'pagina/seguimiento.html', context)


@rol_requerido('CREDITO', 'ADMIN', 'VENDEDOR')
def detalle_nota(request, nota_id):
    """Vista de detalle de una nota con opciones para cambiar estado y agregar comentarios"""
    from .models import NotaCredito, HistorialNota, Aclaracion
    from .forms import CambiarEstadoForm, AgregarComentarioForm
    from django.http import Http404
    
    try:
        nota = NotaCredito.objects.select_related('ruta', 'usuario').get(pk=nota_id)
    except NotaCredito.DoesNotExist:
        raise Http404("La nota no existe")
    
    # Verificar permisos: VENDEDOR solo puede ver notas de sus rutas
    if request.user.rol == 'VENDEDOR':
        if nota.ruta not in request.user.rutas.all():
            messages.error(request, 'No tienes permiso para ver esta nota.')
            return redirect('seguimiento')
    
    # Obtener historial y comentarios
    historial = HistorialNota.objects.filter(nota=nota).select_related('usuario').order_by('-created_at')[:20]
    aclaraciones = Aclaracion.objects.filter(nota=nota).select_related('usuario').order_by('-created_at')
    
    # Si es CREDITO o ADMIN, marcar todos los comentarios como leídos al ver el detalle
    if request.user.rol in ['CREDITO', 'ADMIN']:
        Aclaracion.objects.filter(nota=nota, leida=False).update(leida=True)
        # También marcar las alertas relacionadas como leídas
        from .models import Alerta
        Alerta.objects.filter(nota=nota, tipo='NUEVO_COMENTARIO', usuario_asignado=request.user, leida=False).update(leida=True)
    
    # Formularios
    form_cambiar_estado = CambiarEstadoForm(initial={'estado': nota.estado})
    form_comentario = AgregarComentarioForm()
    
    # Procesar cambio de estado (solo CREDITO y ADMIN)
    if request.method == 'POST' and 'cambiar_estado' in request.POST:
        # Validar que solo CREDITO y ADMIN puedan cambiar estados
        if request.user.rol == 'VENDEDOR':
            messages.error(request, 'No tienes permiso para cambiar el estado de las notas. Solo puedes agregar comentarios.')
            return redirect('detalle_nota', nota_id=nota.id)
        
        form_cambiar_estado = CambiarEstadoForm(request.POST)
        if form_cambiar_estado.is_valid():
            nuevo_estado = form_cambiar_estado.cleaned_data['estado']
            comentario = form_cambiar_estado.cleaned_data['comentario']
            
                # Si el estado cambió, registrar en historial
            if nota.estado != nuevo_estado:
                estado_anterior = nota.get_estado_display()
                nota.estado = nuevo_estado
                
                # Si se marca como resuelta, establecer fecha de resolución y desactivar requiere_atencion
                if nuevo_estado == 'RESUELTA':
                    if not nota.fecha_resolucion:
                        nota.fecha_resolucion = datetime.now()
                    nota.resuelta_automaticamente = False
                    nota.requiere_atencion = False  # Automáticamente desactivar al resolver
                elif nuevo_estado != 'RESUELTA':
                    nota.fecha_resolucion = None
                
                nota.save()
                
                # Registrar en historial
                HistorialNota.objects.create(
                    nota=nota,
                    usuario=request.user,
                    campo_modificado='Estado',
                    valor_anterior=estado_anterior,
                    valor_nuevo=nota.get_estado_display(),
                    observacion=comentario
                )
                
                messages.success(request, f'Estado cambiado a {nota.get_estado_display()} exitosamente.')
            else:
                # Si el estado no cambió pero hay comentario, agregarlo como aclaración
                if comentario:
                    Aclaracion.objects.create(
                        nota=nota,
                        usuario=request.user,
                        comentario=comentario,
                        tipo='COMENTARIO',
                        leida=False
                    )
                    # Activar requiere_atencion cuando se agrega un comentario
                    if not nota.requiere_atencion:
                        nota.requiere_atencion = True
                        nota.save()
                    messages.success(request, 'Comentario agregado exitosamente.')
            
            return redirect('detalle_nota', nota_id=nota.id)
    
    # Procesar agregar comentario
    if request.method == 'POST' and 'agregar_comentario' in request.POST:
        form_comentario = AgregarComentarioForm(request.POST)
        if form_comentario.is_valid():
            from .models import Alerta
            # Crear el comentario
            aclaracion = Aclaracion.objects.create(
                nota=nota,
                usuario=request.user,
                comentario=form_comentario.cleaned_data['comentario'],
                tipo=form_comentario.cleaned_data['tipo'],
                leida=False  # Por defecto no leído
            )
            
            # Activar requiere_atencion cuando se agrega un comentario (cualquier usuario)
            if not nota.requiere_atencion:
                nota.requiere_atencion = True
                nota.save()
            
            messages.success(request, 'Comentario agregado exitosamente.')
            return redirect('detalle_nota', nota_id=nota.id)
    
    context = {
        'titulo': f'Detalle de Nota - {nota.serie_folio}',
        'nota': nota,
        'historial': historial,
        'aclaraciones': aclaraciones,
        'form_cambiar_estado': form_cambiar_estado,
        'form_comentario': form_comentario,
    }
    return render(request, 'pagina/detalle_nota.html', context)


@rol_requerido('ADMIN')
def importar_reporte(request):
    """Importar Reporte CSV, TSV o Excel"""
    from .forms import ImportarReporteForm
    from .models import NotaCredito, Ruta, Importacion
    from .utils import (
        detectar_formato, leer_archivo_csv, leer_archivo_excel,
        procesar_datos, convertir_fecha, convertir_decimal
    )
    from django.db import transaction
    from datetime import date
    
    if request.method == 'POST':
        form = ImportarReporteForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data['archivo']
            nombre_archivo = archivo.name
            formato = detectar_formato(archivo)
            
            # Inicializar variables para el historial
            registros_nuevos = 0
            registros_actualizados = 0
            registros_resueltos = 0
            registros_desaparecidos = 0  # Notas que desaparecieron del reporte
            registros_reaparecidos = 0  # Notas que reaparecieron después de estar resueltas
            total_registros = 0
            errores_procesamiento = []
            estado_importacion = 'FALLIDA'
            observaciones = ''
            
            if not formato:
                mensaje_error = 'Formato de archivo no reconocido.'
                messages.error(request, mensaje_error)
                # Registrar en historial
                Importacion.objects.create(
                    usuario=request.user,
                    nombre_archivo=nombre_archivo,
                    total_registros=0,
                    registros_nuevos=0,
                    registros_actualizados=0,
                    registros_resueltos=0,
                    estado='FALLIDA',
                    observaciones=mensaje_error
                )
                return render(request, 'pagina/importar_reporte.html', {'form': form})
            
            try:
                # Inicializar progreso en la sesión
                request.session['importacion_progreso'] = {
                    'porcentaje': 0,
                    'etapa': 'Iniciando...',
                    'detalle': 'Preparando la importación',
                    'total_registros': 0,
                    'registros_procesados': 0,
                    'completado': False,
                    'error': False
                }
                request.session.save()
                
                # Leer el archivo según su formato
                try:
                    # Actualizar progreso: Validando archivo
                    request.session['importacion_progreso'] = {
                        'porcentaje': 5,
                        'etapa': 'Validando archivo...',
                        'detalle': 'Verificando formato y tamaño del archivo',
                        'total_registros': 0,
                        'registros_procesados': 0,
                        'completado': False,
                        'error': False
                    }
                    request.session.save()
                    
                    if formato == 'csv':
                        datos = leer_archivo_csv(archivo, delimitador=',')
                    elif formato == 'tsv':
                        datos = leer_archivo_csv(archivo, delimitador='\t')
                    elif formato == 'excel':
                        datos = leer_archivo_excel(archivo)
                    else:
                        mensaje_error = 'Formato no soportado.'
                        messages.error(request, mensaje_error)
                        Importacion.objects.create(
                            usuario=request.user,
                            nombre_archivo=nombre_archivo,
                            total_registros=0,
                            registros_nuevos=0,
                            registros_actualizados=0,
                            registros_resueltos=0,
                            estado='FALLIDA',
                            observaciones=mensaje_error
                        )
                        return render(request, 'pagina/importar_reporte.html', {'form': form})
                except Exception as e:
                    mensaje_error = f'Error al leer el archivo: {str(e)}'
                    logger.error(mensaje_error)
                    messages.error(request, mensaje_error)
                    Importacion.objects.create(
                        usuario=request.user,
                        nombre_archivo=nombre_archivo,
                        total_registros=0,
                        registros_nuevos=0,
                        registros_actualizados=0,
                        registros_resueltos=0,
                        estado='FALLIDA',
                        observaciones=mensaje_error
                    )
                    return render(request, 'pagina/importar_reporte.html', {'form': form})
                
                if not datos:
                    mensaje_error = 'El archivo está vacío o no contiene datos válidos.'
                    messages.error(request, mensaje_error)
                    Importacion.objects.create(
                        usuario=request.user,
                        nombre_archivo=nombre_archivo,
                        total_registros=0,
                        registros_nuevos=0,
                        registros_actualizados=0,
                        registros_resueltos=0,
                        estado='FALLIDA',
                        observaciones=mensaje_error
                    )
                    return render(request, 'pagina/importar_reporte.html', {'form': form})
                
                total_registros = len(datos)
                
                # Actualizar progreso: Leyendo archivo
                request.session['importacion_progreso'] = {
                    'porcentaje': 15,
                    'etapa': 'Leyendo archivo...',
                    'detalle': f'Archivo leído: {total_registros} registros encontrados',
                    'total_registros': total_registros,
                    'registros_procesados': 0,
                    'completado': False,
                    'error': False
                }
                request.session.save()
                
                # Mapeo de columnas esperadas (basado en formato del punto de venta)
                mapeo_columnas = {
                    'empresa': ['empresa', 'company', 'compania', 'emp'],
                    'serie_folio': ['serie/folio', 'serie_folio', 'serie folio', 'serie-folio', 'folio', 'serie', 'numero', 'no', 'nro'],
                    'fecha_nota': ['fecha nota', 'fecha_nota', 'fecha', 'fecha_nota_credito', 'fecha de nota'],
                    'fecha_corriente': ['fecha corriente', 'fecha_corriente', 'fecha actual'],  # Columna adicional del PV
                    'cliente': ['cliente', 'customer', 'cliente_nombre', 'nombre_cliente'],
                    'ruta_codigo': ['rutas', 'ruta', 'ruta_codigo', 'codigo_ruta', 'ruta codigo', 'codigo de ruta'],
                    'usuario_vendedor_pv': ['usuario/vendedor', 'usuario_vendedor', 'usuario_vendedor_pv', 'vendedor_pv'],  # Valor interno del PV (ej: PERSONA_1, PERSONA_2)
                    'monto': ['monto', 'amount', 'total', 'importe', 'valor'],
                    'abono': ['abono', 'payment', 'pago', 'pagado'],
                    'saldo': ['saldo', 'balance', 'restante'],
                    'dias': ['dias', 'days', 'días'],  # Columna adicional del PV
                }
                
                # Actualizar progreso: Procesando datos
                request.session['importacion_progreso'] = {
                    'porcentaje': 25,
                    'etapa': 'Procesando datos...',
                    'detalle': 'Validando y mapeando columnas',
                    'total_registros': total_registros,
                    'registros_procesados': 0,
                    'completado': False,
                    'error': False
                }
                request.session.save()
                
                # Procesar datos
                try:
                    registros, errores_parseo = procesar_datos(datos, mapeo_columnas)
                except Exception as e:
                    mensaje_error = f'Error al procesar los datos: {str(e)}'
                    logger.error(mensaje_error)
                    messages.error(request, mensaje_error)
                    Importacion.objects.create(
                        usuario=request.user,
                        nombre_archivo=nombre_archivo,
                        total_registros=total_registros,
                        registros_nuevos=0,
                        registros_actualizados=0,
                        registros_resueltos=0,
                        estado='FALLIDA',
                        observaciones=mensaje_error
                    )
                    return render(request, 'pagina/importar_reporte.html', {'form': form})
                
                if not registros:
                    mensaje_error = 'No se pudieron procesar los datos del archivo.'
                    if errores_parseo:
                        mensaje_error += f' Errores: {"; ".join(errores_parseo[:10])}'
                    messages.error(request, mensaje_error)
                    Importacion.objects.create(
                        usuario=request.user,
                        nombre_archivo=nombre_archivo,
                        total_registros=total_registros,
                        registros_nuevos=0,
                        registros_actualizados=0,
                        registros_resueltos=0,
                        estado='FALLIDA',
                        observaciones=mensaje_error
                    )
                    return render(request, 'pagina/importar_reporte.html', {'form': form})
                
                # OPTIMIZACIÓN 1: Precargar todas las rutas activas en memoria
                logger.info("Precargando rutas activas en memoria...")
                rutas_activas = {ruta.codigo.upper(): ruta for ruta in Ruta.objects.filter(activa=True)}
                logger.info(f"Rutas precargadas: {len(rutas_activas)}")
                
                # OPTIMIZACIÓN 2: Precargar todas las notas existentes que podrían ser actualizadas
                # Extraer todas las combinaciones empresa+serie_folio del reporte primero
                empresas_serie_folio = set()
                for registro in registros:
                    empresa = registro.get('empresa', '').strip()
                    serie_folio = registro.get('serie_folio', '').strip()
                    if empresa and serie_folio:
                        empresas_serie_folio.add((empresa, serie_folio))
                
                logger.info(f"Buscando {len(empresas_serie_folio)} notas existentes...")
                notas_existentes_dict = {}
                if empresas_serie_folio:
                    # Crear consultas optimizadas por empresa
                    empresas_unicas = {emp for emp, _ in empresas_serie_folio}
                    for empresa in empresas_unicas:
                        serie_folios_empresa = {sf for emp, sf in empresas_serie_folio if emp == empresa}
                        notas_empresa = NotaCredito.objects.filter(
                            empresa=empresa,
                            serie_folio__in=serie_folios_empresa
                        )
                        for nota in notas_empresa:
                            notas_existentes_dict[(nota.empresa, nota.serie_folio)] = nota
                logger.info(f"Notas existentes encontradas: {len(notas_existentes_dict)}")
                
                # Actualizar progreso: Preparando procesamiento
                request.session['importacion_progreso'] = {
                    'porcentaje': 40,
                    'etapa': 'Preparando procesamiento...',
                    'detalle': 'Datos precargados, iniciando procesamiento de registros',
                    'total_registros': len(registros),
                    'registros_procesados': 0,
                    'completado': False,
                    'error': False
                }
                request.session.save()
                
                # Pequeño delay para asegurar que el frontend pueda capturar el progreso inicial
                time.sleep(0.1)  # 100ms de delay
                
                # Set para almacenar las combinaciones empresa+serie_folio del reporte actual
                notas_en_reporte = set()
                
                # Listas para bulk operations
                notas_a_crear = []
                notas_a_actualizar = []
                
                # PROCESAR EN LOTES para permitir actualizaciones de progreso en tiempo real
                TAMANO_LOTE = 150  # Procesar 150 registros por lote
                total_registros = len(registros)
                num_lotes = (total_registros + TAMANO_LOTE - 1) // TAMANO_LOTE  # Redondear hacia arriba
                
                for num_lote in range(num_lotes):
                    inicio_lote = num_lote * TAMANO_LOTE
                    fin_lote = min(inicio_lote + TAMANO_LOTE, total_registros)
                    lote_registros = registros[inicio_lote:fin_lote]
                    
                    # Procesar el lote dentro de una transacción
                    with transaction.atomic():
                        for idx_relativo, registro in enumerate(lote_registros, start=1):
                            idx_absoluto = inicio_lote + idx_relativo
                            
                            try:
                                # Validar campos requeridos
                                empresa = registro.get('empresa', '').strip()
                                serie_folio = registro.get('serie_folio', '').strip()
                                
                                if not empresa or not serie_folio:
                                    errores_procesamiento.append(f"Fila {idx_absoluto + 1}: Empresa y Serie-Folio son requeridos")
                                    continue
                                
                                # Agregar a set de notas en el reporte
                                notas_en_reporte.add((empresa, serie_folio))
                                
                                # OPTIMIZACIÓN 1: Usar diccionario de rutas precargadas
                                ruta_codigo = registro.get('ruta_codigo', '').strip()
                                if not ruta_codigo:
                                    errores_procesamiento.append(f"Fila {idx_absoluto + 1}: Código de ruta es requerido")
                                    continue
                                
                                ruta = rutas_activas.get(ruta_codigo.upper())
                                if not ruta:
                                    errores_procesamiento.append(f"Fila {idx_absoluto + 1}: Ruta '{ruta_codigo}' no existe o está inactiva")
                                    continue
                                
                                # Convertir fechas
                                try:
                                    fecha_nota = convertir_fecha(registro.get('fecha_nota', ''))
                                    if not fecha_nota:
                                        fecha_nota = date.today()
                                except:
                                    fecha_nota = date.today()
                                
                                # Obtener fecha corriente del registro (si viene del PV) o usar fecha actual
                                try:
                                    fecha_corriente = convertir_fecha(registro.get('fecha_corriente', ''))
                                    if not fecha_corriente:
                                        fecha_corriente = date.today()
                                except:
                                    fecha_corriente = date.today()
                                
                                # Convertir montos
                                monto = convertir_decimal(registro.get('monto', '0'))
                                abono = convertir_decimal(registro.get('abono', '0'))
                                saldo = monto - abono
                                
                                if monto <= 0:
                                    errores_procesamiento.append(f"Fila {idx_absoluto + 1}: El monto debe ser mayor a cero")
                                    continue
                                
                                # Obtener cliente
                                cliente = registro.get('cliente', '').strip()
                                if not cliente:
                                    cliente = 'Sin nombre'
                                
                                # Obtener usuario/vendedor del punto de venta (valor interno, no relacionado con usuario_id)
                                # Este es un valor interno del PV como "PERSONA_1", "PERSONA_2", etc.
                                # NO tiene relación con la tabla usuarios, solo se guarda como texto informativo
                                usuario_vendedor_pv = registro.get('usuario_vendedor_pv', '').strip()
                                if not usuario_vendedor_pv:
                                    usuario_vendedor_pv = None
                                
                                # OPTIMIZACIÓN 2: Usar diccionario de notas precargadas
                                nota_key = (empresa, serie_folio)
                                nota_existente = notas_existentes_dict.get(nota_key)
                                
                                if nota_existente:
                                    # Verificar si la nota estaba resuelta o cancelada y ahora reaparece
                                    estado_anterior = nota_existente.estado
                                    estaba_resuelta = nota_existente.estado == 'RESUELTA'
                                    estaba_cancelada = nota_existente.estado == 'CANCELADA'
                                    
                                    # Actualizar nota existente
                                    nota_existente.fecha_nota = fecha_nota
                                    nota_existente.fecha_corriente = fecha_corriente
                                    nota_existente.cliente = cliente
                                    nota_existente.ruta = ruta
                                    nota_existente.usuario_vendedor_pv = usuario_vendedor_pv
                                    nota_existente.monto = monto
                                    nota_existente.abono = abono
                                    nota_existente.saldo = saldo
                                    
                                    # Si la nota reaparece en el reporte, determinar su nuevo estado
                                    if saldo <= 0:
                                        # Si el saldo es 0, siempre está resuelta
                                        nota_existente.estado = 'RESUELTA'
                                        nota_existente.resuelta_automaticamente = True
                                        nota_existente.fecha_resolucion = datetime.now()
                                        nota_existente.requiere_atencion = False  # Desactivar al resolver
                                        registros_resueltos += 1
                                    else:
                                        # Si tiene saldo > 0 y estaba resuelta o cancelada, reactivarla
                                        if estaba_resuelta or estaba_cancelada:
                                            nota_existente.estado = 'PENDIENTE'
                                            nota_existente.resuelta_automaticamente = False
                                            nota_existente.fecha_resolucion = None
                                            registros_reaparecidos += 1
                                        # Si ya estaba pendiente, mantener pendiente
                                        elif nota_existente.estado == 'PENDIENTE':
                                            # Ya está pendiente, no hacer nada
                                            pass
                                    
                                    # Agregar a lista para bulk_update
                                    notas_a_actualizar.append(nota_existente)
                                    registros_actualizados += 1
                                    
                                else:
                                    # Crear nueva nota
                                    nueva_nota = NotaCredito(
                                        empresa=empresa,
                                        serie_folio=serie_folio,
                                        fecha_nota=fecha_nota,
                                        cliente=cliente,
                                        ruta=ruta,
                                        usuario_vendedor_pv=usuario_vendedor_pv,
                                        monto=monto,
                                        abono=abono,
                                        saldo=saldo,
                                        estado='PENDIENTE',
                                        fecha_corriente=fecha_corriente
                                    )
                                    
                                    # Si el saldo es 0 desde el inicio, marcar como resuelta
                                    if saldo <= 0:
                                        nueva_nota.estado = 'RESUELTA'
                                        nueva_nota.resuelta_automaticamente = True
                                        nueva_nota.fecha_resolucion = datetime.now()
                                        nueva_nota.requiere_atencion = False  # Desactivar al resolver
                                        registros_resueltos += 1
                                    
                                    # Agregar a lista para bulk_create
                                    notas_a_crear.append(nueva_nota)
                                    registros_nuevos += 1
                                    
                                    # Agregar al diccionario para evitar duplicados
                                    notas_existentes_dict[nota_key] = nueva_nota
                                
                            except Exception as e:
                                logger.error(f"Error procesando registro {idx_absoluto}: {str(e)}")
                                errores_procesamiento.append(f"Fila {idx_absoluto + 1}: {str(e)}")
                    
                    # Guardar el lote procesado
                    if notas_a_crear:
                        NotaCredito.objects.bulk_create(notas_a_crear, batch_size=500)
                        notas_a_crear = []  # Limpiar lista para el siguiente lote
                    if notas_a_actualizar:
                        NotaCredito.objects.bulk_update(
                            notas_a_actualizar,
                            ['fecha_nota', 'fecha_corriente', 'cliente', 'ruta', 'usuario_vendedor_pv',
                             'monto', 'abono', 'saldo', 'estado', 'resuelta_automaticamente',
                             'fecha_resolucion', 'requiere_atencion'],
                            batch_size=500
                        )
                        notas_a_actualizar = []  # Limpiar lista para el siguiente lote
                    
                    # ACTUALIZAR PROGRESO FUERA DE LA TRANSACCIÓN para que se refleje en tiempo real
                    porcentaje = 40 + int((fin_lote / total_registros) * 40)  # 40% a 80%
                    request.session['importacion_progreso'] = {
                        'porcentaje': porcentaje,
                        'etapa': 'Procesando registros...',
                        'detalle': f'Procesando registro {fin_lote} de {total_registros}',
                        'total_registros': total_registros,
                        'registros_procesados': fin_lote,
                        'completado': False,
                        'error': False
                    }
                    request.session.save()
                    logger.info(f"Lote {num_lote + 1}/{num_lotes} completado: {fin_lote}/{total_registros} registros - Progreso: {porcentaje}%")
                    
                    # Pequeño delay para permitir que el frontend capture el progreso
                    time.sleep(0.05)  # 50ms de delay entre lotes
                
                # OPTIMIZACIÓN 4: Guardar cualquier nota restante
                logger.info(f"Guardando notas restantes: {len(notas_a_crear)} nuevas, {len(notas_a_actualizar)} actualizaciones...")
                if notas_a_crear:
                    NotaCredito.objects.bulk_create(notas_a_crear, batch_size=500)
                if notas_a_actualizar:
                    NotaCredito.objects.bulk_update(
                        notas_a_actualizar,
                        ['fecha_nota', 'fecha_corriente', 'cliente', 'ruta', 'usuario_vendedor_pv',
                         'monto', 'abono', 'saldo', 'estado', 'resuelta_automaticamente',
                         'fecha_resolucion', 'requiere_atencion'],
                        batch_size=500
                    )
                logger.info("Bulk operations completadas")
                    
                # Actualizar progreso: Detectando cambios
                request.session['importacion_progreso'] = {
                    'porcentaje': 80,
                    'etapa': 'Detectando cambios...',
                    'detalle': 'Analizando notas desaparecidas y reaparecidas',
                    'total_registros': len(registros),
                    'registros_procesados': len(registros),
                    'completado': False,
                    'error': False
                }
                request.session.save()
                
                # Detectar notas desaparecidas: Notas PENDIENTE que no aparecen en el nuevo reporte
                # IMPORTANTE: Solo revisar notas de las empresas que están en el reporte importado
                # Extraer las empresas únicas del reporte
                empresas_en_reporte = set()
                for empresa, serie_folio in notas_en_reporte:
                    empresas_en_reporte.add(empresa)
                
                # OPTIMIZACIÓN 5: Usar bulk_update para notas desaparecidas
                with transaction.atomic():
                    if empresas_en_reporte:
                        notas_pendientes = NotaCredito.objects.filter(
                            estado='PENDIENTE',
                            empresa__in=empresas_en_reporte
                        )
                        notas_a_resolver = []
                        for nota in notas_pendientes:
                            if (nota.empresa, nota.serie_folio) not in notas_en_reporte:
                                # La nota desapareció del reporte, se considera resuelta
                                nota.estado = 'RESUELTA'
                                nota.resuelta_automaticamente = True
                                nota.fecha_resolucion = datetime.now()
                                nota.requiere_atencion = False  # Desactivar al resolver automáticamente
                                notas_a_resolver.append(nota)
                                registros_desaparecidos += 1
                        
                        if notas_a_resolver:
                            NotaCredito.objects.bulk_update(
                                notas_a_resolver,
                                ['estado', 'resuelta_automaticamente', 'fecha_resolucion', 'requiere_atencion'],
                                batch_size=500
                            )
                
                # Actualizar progreso: Finalizando
                request.session['importacion_progreso'] = {
                    'porcentaje': 90,
                    'etapa': 'Finalizando...',
                    'detalle': 'Guardando resultados y generando reporte',
                    'total_registros': len(registros),
                    'registros_procesados': len(registros),
                    'completado': False,
                    'error': False
                }
                request.session.save()
                
                # Determinar estado de importación
                estado_importacion = 'COMPLETADA'
                if errores_procesamiento:
                    estado_importacion = 'PARCIAL' if (registros_nuevos + registros_actualizados) > 0 else 'FALLIDA'
                    
                    # Preparar observaciones
                    observaciones_parts = []
                    if errores_procesamiento:
                        observaciones_parts.append(f"Total de errores: {len(errores_procesamiento)}. ")
                        observaciones_parts.append("; ".join(errores_procesamiento[:20]))  # Mostrar hasta 20 errores
                        if len(errores_procesamiento) > 20:
                            observaciones_parts.append(f" ... y {len(errores_procesamiento) - 20} errores más.")
                    
                    if registros_desaparecidos > 0:
                        observaciones_parts.append(f"Notas desaparecidas (resueltas automáticamente): {registros_desaparecidos}.")
                    
                    if registros_reaparecidos > 0:
                        observaciones_parts.append(f"Notas reaparecidas (reactivadas): {registros_reaparecidos}.")
                    
                    observaciones = " ".join(observaciones_parts)
                    
                    # Crear registro de importación (SIEMPRE se crea, incluso con errores)
                    importacion = Importacion.objects.create(
                        usuario=request.user,
                        nombre_archivo=nombre_archivo,
                        total_registros=len(registros),
                        registros_nuevos=registros_nuevos,
                        registros_actualizados=registros_actualizados,
                        registros_resueltos=registros_resueltos,
                        estado=estado_importacion,
                        observaciones=observaciones
                    )
                
                # Mensajes de éxito
                mensaje = f"Importación completada: {registros_nuevos} nuevos, {registros_actualizados} actualizados"
                if registros_resueltos > 0:
                    mensaje += f", {registros_resueltos} resueltos automáticamente"
                if registros_desaparecidos > 0:
                    mensaje += f", {registros_desaparecidos} desaparecidas (resueltas)"
                if registros_reaparecidos > 0:
                    mensaje += f", {registros_reaparecidos} reaparecidas (reactivadas)"
                messages.success(request, mensaje)
                
                if errores_procesamiento:
                    messages.warning(request, f"Se encontraron {len(errores_procesamiento)} errores durante la importación.")
                
                # Marcar progreso como completado antes de limpiarlo
                request.session['importacion_progreso'] = {
                    'porcentaje': 100,
                    'etapa': '¡Completado!',
                    'detalle': 'La importación se ha completado exitosamente',
                    'total_registros': len(registros),
                    'registros_procesados': len(registros),
                    'completado': True,
                    'error': False
                }
                request.session.save()
                
                # Si es una petición AJAX, responder con JSON
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    logger.info(f"Importación completada vía AJAX: {registros_nuevos} nuevos, {registros_actualizados} actualizados")
                    return JsonResponse({
                        'success': True,
                        'message': 'Importación completada exitosamente',
                        'registros_nuevos': registros_nuevos,
                        'registros_actualizados': registros_actualizados,
                        'registros_resueltos': registros_resueltos,
                        'registros_desaparecidos': registros_desaparecidos,
                        'registros_reaparecidos': registros_reaparecidos,
                        'total_registros': len(registros),
                        'errores': len(errores_procesamiento)
                    })
                
                return redirect('importar_reporte')
                
            except Exception as e:
                error_msg = f'Error al procesar el archivo: {str(e)}'
                logger.error(f"Error en importación: {error_msg}")
                messages.error(request, error_msg)
                # Registrar error en historial
                try:
                    Importacion.objects.create(
                        usuario=request.user,
                        nombre_archivo=nombre_archivo if 'nombre_archivo' in locals() else 'archivo_desconocido',
                        total_registros=total_registros if 'total_registros' in locals() else 0,
                        registros_nuevos=0,
                        registros_actualizados=0,
                        registros_resueltos=0,
                        estado='FALLIDA',
                        observaciones=error_msg
                    )
                except Exception as db_error:
                    logger.error(f"Error al guardar historial de importación: {str(db_error)}")
    else:
        form = ImportarReporteForm()
    
    # Obtener últimas importaciones
    ultimas_importaciones = Importacion.objects.select_related('usuario').order_by('-created_at')[:10]
    
    context = {
        'form': form,
        'ultimas_importaciones': ultimas_importaciones,
    }
    return render(request, 'pagina/importar_reporte.html', context)


@rol_requerido('ADMIN')
def ultimas_importaciones_ajax(request):
    """Vista AJAX para obtener las últimas importaciones (solo el fragmento HTML)"""
    from .models import Importacion
    
    # Obtener últimas importaciones
    ultimas_importaciones = Importacion.objects.select_related('usuario').order_by('-created_at')[:10]
    
    context = {
        'ultimas_importaciones': ultimas_importaciones,
    }
    return render(request, 'pagina/ultimas_importaciones_fragment.html', context)


@rol_requerido('ADMIN')
def importacion_progreso_ajax(request):
    """Vista AJAX para obtener el progreso actual de la importación"""
    from django.http import JsonResponse
    
    progreso = request.session.get('importacion_progreso', {
        'porcentaje': 0,
        'etapa': 'No hay importación en curso',
        'detalle': '',
        'total_registros': 0,
        'registros_procesados': 0,
        'completado': False,
        'error': False
    })
    
    # Si está completado, limpiar después de devolver la respuesta
    # (el frontend ya habrá leído el estado completado)
    if progreso.get('completado', False):
        # Limpiar inmediatamente después de devolver
        if 'importacion_progreso' in request.session:
            del request.session['importacion_progreso']
            request.session.save()
    
    return JsonResponse(progreso)


@rol_requerido('ADMIN')
def descargar_archivo_muestra(request):
    """Generar y descargar archivo de muestra para importación"""
    import csv
    import io
    from datetime import date, timedelta
    from django.http import HttpResponse
    from .models import Ruta
    
    formato = request.GET.get('formato', 'csv').lower()
    
    # Obtener rutas activas de la base de datos
    rutas_activas = Ruta.objects.filter(activa=True).order_by('codigo')[:5]
    codigos_rutas = [ruta.codigo for ruta in rutas_activas]
    
    # Si no hay rutas, usar códigos de ejemplo
    if not codigos_rutas:
        codigos_rutas = ['DR201', 'DR202', 'DR203', 'DR204', 'DR201']
    
    # Asegurar que tengamos al menos 5 códigos (repetir si es necesario)
    while len(codigos_rutas) < 5:
        codigos_rutas.append(codigos_rutas[0] if codigos_rutas else 'DR201')
    
    # Valores de ejemplo para usuario/vendedor del punto de venta (valores internos del PV)
    valores_pv = ['PERSONA_1', 'PERSONA_2', 'PERSONA_1', 'PERSONA_2', 'PERSONA_1']
    
    # Datos de ejemplo (formato del punto de venta)
    fecha_corriente = date.today().strftime('%Y-%m-%d')
    datos_muestra = [
        {
            'FECHA NOTA': (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'),
            'FECHA CORRIENTE': fecha_corriente,
            'SERIE/FOLIO': 'NC-001-2024',
            'CLIENTE': 'Cliente Ejemplo 1',
            'RUTAS': codigos_rutas[0],
            'USUARIO/VENDEDOR': valores_pv[0],
            'MONTO': '5000.00',
            'ABONO': '2000.00',
            'SALDO': '3000.00',
            'DIAS': '30',
            'EMPRESA': 'DISTRIBUIDORA'
        },
        {
            'FECHA NOTA': (date.today() - timedelta(days=25)).strftime('%Y-%m-%d'),
            'FECHA CORRIENTE': fecha_corriente,
            'SERIE/FOLIO': 'NC-002-2024',
            'CLIENTE': 'Cliente Ejemplo 2',
            'RUTAS': codigos_rutas[1] if len(codigos_rutas) > 1 else codigos_rutas[0],
            'USUARIO/VENDEDOR': valores_pv[1],
            'MONTO': '7500.50',
            'ABONO': '0.00',
            'SALDO': '7500.50',
            'DIAS': '25',
            'EMPRESA': 'RODRIGO'
        },
        {
            'FECHA NOTA': (date.today() - timedelta(days=20)).strftime('%Y-%m-%d'),
            'FECHA CORRIENTE': fecha_corriente,
            'SERIE/FOLIO': 'NC-003-2024',
            'CLIENTE': 'Cliente Ejemplo 3',
            'RUTAS': codigos_rutas[2] if len(codigos_rutas) > 2 else codigos_rutas[0],
            'USUARIO/VENDEDOR': valores_pv[2],
            'MONTO': '3200.75',
            'ABONO': '3200.75',
            'SALDO': '0.00',
            'DIAS': '20',
            'EMPRESA': 'DISTRIBUIDORA'
        },
        {
            'FECHA NOTA': (date.today() - timedelta(days=15)).strftime('%Y-%m-%d'),
            'FECHA CORRIENTE': fecha_corriente,
            'SERIE/FOLIO': 'NC-004-2024',
            'CLIENTE': 'Cliente Ejemplo 4',
            'RUTAS': codigos_rutas[3] if len(codigos_rutas) > 3 else codigos_rutas[0],
            'USUARIO/VENDEDOR': valores_pv[3],
            'MONTO': '12000.00',
            'ABONO': '5000.00',
            'SALDO': '7000.00',
            'DIAS': '15',
            'EMPRESA': 'RODRIGO'
        },
        {
            'FECHA NOTA': (date.today() - timedelta(days=10)).strftime('%Y-%m-%d'),
            'FECHA CORRIENTE': fecha_corriente,
            'SERIE/FOLIO': 'NC-005-2024',
            'CLIENTE': 'Cliente Ejemplo 5',
            'RUTAS': codigos_rutas[4] if len(codigos_rutas) > 4 else codigos_rutas[0],
            'USUARIO/VENDEDOR': valores_pv[4],
            'MONTO': '8500.25',
            'ABONO': '3000.00',
            'SALDO': '5500.25',
            'DIAS': '10',
            'EMPRESA': 'DISTRIBUIDORA'
        }
    ]
    
    # Encabezados (formato del punto de venta)
    encabezados = ['FECHA NOTA', 'FECHA CORRIENTE', 'SERIE/FOLIO', 'CLIENTE', 'RUTAS', 'USUARIO/VENDEDOR', 'MONTO', 'ABONO', 'SALDO', 'DIAS', 'EMPRESA']
    
    if formato == 'excel' or formato == 'xlsx':
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Notas de Crédito"
            
            # Agregar encabezados
            for col, encabezado in enumerate(encabezados, start=1):
                cell = ws.cell(row=1, column=col, value=encabezado)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="4CA150", end_color="4CA150", fill_type="solid")
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            
            # Agregar datos
            for row, dato in enumerate(datos_muestra, start=2):
                ws.cell(row=row, column=1, value=dato['FECHA NOTA'])
                ws.cell(row=row, column=2, value=dato['FECHA CORRIENTE'])
                ws.cell(row=row, column=3, value=dato['SERIE/FOLIO'])
                ws.cell(row=row, column=4, value=dato['CLIENTE'])
                ws.cell(row=row, column=5, value=dato['RUTAS'])
                ws.cell(row=row, column=6, value=dato['USUARIO/VENDEDOR'])
                ws.cell(row=row, column=7, value=float(dato['MONTO']))
                ws.cell(row=row, column=8, value=float(dato['ABONO']))
                ws.cell(row=row, column=9, value=float(dato['SALDO']))
                ws.cell(row=row, column=10, value=int(dato['DIAS']))
                ws.cell(row=row, column=11, value=dato['EMPRESA'])
            
            # Ajustar ancho de columnas
            for col in range(1, len(encabezados) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            # Crear respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="muestra_notas_credito.xlsx"'
            
            wb.save(response)
            return response
            
        except ImportError:
            # Si no está openpyxl, generar CSV
            formato = 'csv'
    
    # Generar CSV (por defecto o si Excel falla)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=encabezados)
    writer.writeheader()
    writer.writerows(datos_muestra)
    
    # Preparar contenido con BOM para Excel
    contenido = '\ufeff' + output.getvalue()
    
    response = HttpResponse(contenido, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="muestra_notas_credito.csv"'
    
    return response


@rol_requerido('ADMIN')
def historial_importaciones(request):
    """Historial de Importaciones"""
    from .models import Importacion
    
    # Obtener todas las importaciones ordenadas por fecha (más recientes primero)
    importaciones = Importacion.objects.all().order_by('-created_at')
    
    # Estadísticas
    total_importaciones = importaciones.count()
    completadas = importaciones.filter(estado='COMPLETADA').count()
    parciales = importaciones.filter(estado='PARCIAL').count()
    fallidas = importaciones.filter(estado='FALLIDA').count()
    
    context = {
        'importaciones': importaciones,
        'total_importaciones': total_importaciones,
        'completadas': completadas,
        'parciales': parciales,
        'fallidas': fallidas,
    }
    return render(request, 'pagina/historial_importaciones.html', context)


@rol_requerido('ADMIN')
def usuarios(request):
    """Lista de usuarios"""
    from .models import Usuario
    from .forms import UsuarioForm
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    
    usuarios_list = Usuario.objects.all().order_by('username')
    form = UsuarioForm()  # Inicializar form siempre
    
    if request.method == 'POST':
        logger.info(f'[USUARIOS] ===== PETICIÓN POST RECIBIDA =====')
        logger.info(f'[USUARIOS] Usuario autenticado: {request.user.username} (ID: {request.user.id})')
        logger.info(f'[USUARIOS] Todos los datos POST: {dict(request.POST)}')
        logger.info(f'[USUARIOS] Claves en POST: {list(request.POST.keys())}')
        logger.info(f'[USUARIOS] ¿Contiene "crear"?: {"crear" in request.POST}')
        logger.info(f'[USUARIOS] ¿Contiene "eliminar"?: {"eliminar" in request.POST}')
        logger.info(f'[USUARIOS] ¿Contiene "toggle_activo"?: {"toggle_activo" in request.POST}')
        if 'crear' in request.POST:
            form = UsuarioForm(request.POST)
            if form.is_valid():
                usuario = form.save()
                messages.success(request, f'Usuario "{usuario.username}" creado exitosamente.')
                return redirect('usuarios')
            else:
                messages.error(request, 'Por favor corrige los errores en el formulario.')
        elif 'eliminar' in request.POST:
            logger.info(f'[ELIMINAR_USUARIO] ===== INICIANDO PROCESO DE ELIMINACIÓN =====')
            logger.info(f'[ELIMINAR_USUARIO] Usuario que solicita eliminación: {request.user.username} (ID: {request.user.id})')
            logger.info(f'[ELIMINAR_USUARIO] Método de petición: {request.method}')
            logger.info(f'[ELIMINAR_USUARIO] Datos POST recibidos: {dict(request.POST)}')
            
            usuario_id = request.POST.get('usuario_id')
            logger.info(f'[ELIMINAR_USUARIO] usuario_id obtenido del POST: {usuario_id}')
            logger.info(f'[ELIMINAR_USUARIO] Tipo de usuario_id: {type(usuario_id)}')
            
            if not usuario_id:
                logger.error(f'[ELIMINAR_USUARIO] ERROR: usuario_id es None o vacío')
                messages.error(request, 'No se proporcionó el ID del usuario a eliminar.')
                return redirect('usuarios')
            
            try:
                usuario_id_int = int(usuario_id)
                logger.info(f'[ELIMINAR_USUARIO] usuario_id convertido a int: {usuario_id_int}')
            except (ValueError, TypeError) as e:
                logger.error(f'[ELIMINAR_USUARIO] ERROR al convertir usuario_id a int: {e}')
                messages.error(request, f'ID de usuario inválido: {usuario_id}')
                return redirect('usuarios')
            
            try:
                usuario = Usuario.objects.get(id=usuario_id_int)
                logger.info(f'[ELIMINAR_USUARIO] Usuario encontrado en BD:')
                logger.info(f'[ELIMINAR_USUARIO]   - ID: {usuario.id}')
                logger.info(f'[ELIMINAR_USUARIO]   - Username: {usuario.username}')
                logger.info(f'[ELIMINAR_USUARIO]   - Email: {usuario.email}')
                logger.info(f'[ELIMINAR_USUARIO]   - Rol: {usuario.rol}')
                logger.info(f'[ELIMINAR_USUARIO]   - Activo: {usuario.activo}')
            except Usuario.DoesNotExist:
                logger.error(f'[ELIMINAR_USUARIO] ERROR: Usuario con ID {usuario_id_int} no existe en la BD')
                messages.error(request, f'Usuario con ID {usuario_id_int} no encontrado.')
                return redirect('usuarios')
            except Exception as e:
                logger.error(f'[ELIMINAR_USUARIO] ERROR inesperado al buscar usuario: {type(e).__name__}: {e}')
                messages.error(request, f'Error al buscar usuario: {str(e)}')
                return redirect('usuarios')
            
            # Verificar si es el mismo usuario
            logger.info(f'[ELIMINAR_USUARIO] Comparando usuarios:')
            logger.info(f'[ELIMINAR_USUARIO]   - Usuario a eliminar: {usuario.username} (ID: {usuario.id})')
            logger.info(f'[ELIMINAR_USUARIO]   - Usuario actual: {request.user.username} (ID: {request.user.id})')
            logger.info(f'[ELIMINAR_USUARIO]   - ¿Son iguales?: {usuario == request.user}')
            logger.info(f'[ELIMINAR_USUARIO]   - ¿IDs iguales?: {usuario.id == request.user.id}')
            
            if usuario == request.user:
                logger.warning(f'[ELIMINAR_USUARIO] BLOQUEADO: Intento de eliminar propio usuario')
                messages.error(request, 'No puedes eliminar tu propio usuario.')
                return redirect('usuarios')
            
            # Verificar si el usuario existe antes de eliminar
            try:
                usuario_verificar = Usuario.objects.get(id=usuario_id_int)
                logger.info(f'[ELIMINAR_USUARIO] Verificación pre-eliminación: Usuario existe en BD')
            except Usuario.DoesNotExist:
                logger.error(f'[ELIMINAR_USUARIO] ERROR: Usuario desapareció antes de eliminar')
                messages.error(request, 'El usuario ya no existe en la base de datos.')
                return redirect('usuarios')
            
            # Guardar información antes de eliminar
                username = usuario.username
            usuario_id_backup = usuario.id
            
            logger.info(f'[ELIMINAR_USUARIO] Intentando eliminar usuario: {username} (ID: {usuario_id_backup})')
            
            try:
                usuario.delete()
                logger.info(f'[ELIMINAR_USUARIO] Método delete() ejecutado sin excepciones')
                
                # Verificar que se eliminó correctamente
                try:
                    usuario_verificar = Usuario.objects.get(id=usuario_id_backup)
                    logger.error(f'[ELIMINAR_USUARIO] ERROR: El usuario todavía existe después de delete()')
                    logger.error(f'[ELIMINAR_USUARIO] Usuario encontrado: {usuario_verificar.username}')
                    messages.error(request, f'Error: El usuario "{username}" no se pudo eliminar correctamente.')
                except Usuario.DoesNotExist:
                    logger.info(f'[ELIMINAR_USUARIO] ✓ Usuario eliminado correctamente de la BD')
                    logger.info(f'[ELIMINAR_USUARIO] Verificación post-eliminación: Usuario no existe (correcto)')
                messages.success(request, f'Usuario "{username}" eliminado exitosamente.')
                
            except Exception as e:
                logger.error(f'[ELIMINAR_USUARIO] ERROR al ejecutar delete(): {type(e).__name__}: {e}')
                logger.error(f'[ELIMINAR_USUARIO] Detalles del error: {str(e)}')
                import traceback
                logger.error(f'[ELIMINAR_USUARIO] Traceback completo:\n{traceback.format_exc()}')
                messages.error(request, f'Error al eliminar usuario: {str(e)}')
            
            logger.info(f'[ELIMINAR_USUARIO] ===== FIN DEL PROCESO DE ELIMINACIÓN =====')
            return redirect('usuarios')
        elif 'toggle_activo' in request.POST:
            usuario_id = request.POST.get('usuario_id')
            logger.info(f'[TOGGLE_ACTIVO] Iniciando toggle para usuario_id: {usuario_id}')
            
            usuario = get_object_or_404(Usuario, id=usuario_id)
            logger.info(f'[TOGGLE_ACTIVO] Usuario encontrado: {usuario.username}, Estado actual: {usuario.activo}')
            
            if usuario == request.user:
                logger.warning(f'[TOGGLE_ACTIVO] Intento de desactivar propio usuario: {request.user.username}')
                messages.error(request, 'No puedes desactivar tu propio usuario.')
            else:
                estado_anterior = usuario.activo
                usuario.activo = not usuario.activo
                logger.info(f'[TOGGLE_ACTIVO] Cambiando estado de {estado_anterior} a {usuario.activo}')
                
                usuario.save()
                logger.info(f'[TOGGLE_ACTIVO] Usuario guardado. Verificando estado en BD...')
                
                # Verificar que se guardó correctamente
                usuario_refreshed = Usuario.objects.get(id=usuario.id)
                logger.info(f'[TOGGLE_ACTIVO] Estado después de guardar: {usuario_refreshed.activo}')
                
                if usuario_refreshed.activo != usuario.activo:
                    logger.error(f'[TOGGLE_ACTIVO] ERROR: El estado no se guardó correctamente!')
                
                estado = 'activado' if usuario.activo else 'desactivado'
                messages.success(request, f'Usuario "{usuario.username}" {estado} exitosamente.')
                logger.info(f'[TOGGLE_ACTIVO] Proceso completado exitosamente. Usuario: {usuario.username}, Estado: {estado}')
            
            return redirect('usuarios')
    
    # Refrescar usuarios desde la BD antes de renderizar
    usuarios_list = Usuario.objects.all().order_by('username')
    logger.info(f'[USUARIOS] Refrescando lista de usuarios antes de renderizar')
    for u in usuarios_list:
        logger.debug(f'[USUARIOS] Usuario en contexto: {u.username}, Activo: {u.activo}')
    
    context = {
        'titulo': 'Gestión de Usuarios',
        'usuarios': usuarios_list,
        'form': form,
    }
    logger.info(f'[USUARIOS] Renderizando template con {usuarios_list.count()} usuarios')
    return render(request, 'pagina/usuarios.html', context)


@rol_requerido('ADMIN')
def editar_usuario(request, usuario_id):
    """Editar un usuario específico"""
    from .models import Usuario
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    if request.method == 'POST':
        # Actualizar campos
        usuario.email = request.POST.get('email')
        usuario.nombre_completo = request.POST.get('nombre_completo')
        usuario.rol = request.POST.get('rol')
        usuario.activo = request.POST.get('activo') == 'on'
        
        # Cambiar contraseña si se proporcionó
        password = request.POST.get('password')
        if password:
            usuario.set_password(password)
        
        try:
            usuario.save()
            messages.success(request, f'Usuario "{usuario.username}" actualizado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al actualizar usuario: {str(e)}')
        
        return redirect('usuarios')
    
    context = {
        'titulo': f'Editar Usuario: {usuario.username}',
        'usuario': usuario,
    }
    return render(request, 'pagina/editar_usuario.html', context)


@rol_requerido('ADMIN')
def editar_ruta(request, ruta_id):
    """Editar una ruta específica"""
    from .models import Ruta
    from .forms import RutaForm
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    
    ruta = get_object_or_404(Ruta, id=ruta_id)
    
    if request.method == 'POST':
        # Actualizar campos directamente
        ruta.nombre = request.POST.get('nombre')
        ruta.descripcion = request.POST.get('descripcion', '')
        ruta.activa = request.POST.get('activa') == 'on'
        
        try:
            ruta.save()
            messages.success(request, f'Ruta "{ruta.codigo}" actualizada exitosamente.')
            return redirect('rutas')
        except Exception as e:
            messages.error(request, f'Error al actualizar ruta: {str(e)}')
            return redirect('rutas')
    else:
        form = RutaForm(instance=ruta)
    
    context = {
        'titulo': f'Editar Ruta: {ruta.codigo}',
        'ruta': ruta,
        'form': form,
    }
    return render(request, 'pagina/editar_ruta.html', context)


@rol_requerido('ADMIN')
def asignar_rutas_usuario(request, usuario_id):
    """Asignar rutas a un usuario"""
    from .models import Usuario, UsuarioRuta
    from .forms import AsignarRutasForm
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    if request.method == 'POST':
        form = AsignarRutasForm(request.POST)
        if form.is_valid():
            rutas_seleccionadas = form.cleaned_data['rutas']
            
            # Eliminar asignaciones actuales
            UsuarioRuta.objects.filter(usuario=usuario).delete()
            
            # Crear nuevas asignaciones
            for ruta in rutas_seleccionadas:
                UsuarioRuta.objects.create(usuario=usuario, ruta=ruta)
            
            messages.success(request, f'Rutas asignadas a "{usuario.username}" exitosamente.')
            return redirect('usuarios')
    else:
        # Pre-seleccionar rutas actuales del usuario
        rutas_actuales = UsuarioRuta.objects.filter(usuario=usuario).values_list('ruta_id', flat=True)
        form = AsignarRutasForm(initial={'rutas': rutas_actuales})
    
    # Obtener rutas asignadas actualmente
    rutas_asignadas = UsuarioRuta.objects.filter(usuario=usuario).select_related('ruta')
    
    context = {
        'titulo': f'Asignar Rutas a: {usuario.username}',
        'usuario': usuario,
        'form': form,
        'rutas_asignadas': rutas_asignadas,
    }
    return render(request, 'pagina/asignar_rutas.html', context)


@rol_requerido('ADMIN')
def rutas(request):
    """Gestión de Rutas"""
    from .models import Ruta, UsuarioRuta
    from .forms import RutaForm
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    
    rutas_list = Ruta.objects.all().order_by('codigo')
    form = RutaForm()
    
    if request.method == 'POST':
        logger.info(f'[RUTAS] ===== PETICIÓN POST RECIBIDA =====')
        logger.info(f'[RUTAS] Usuario: {request.user.username}')
        logger.info(f'[RUTAS] Datos POST: {dict(request.POST)}')
        
        if 'crear' in request.POST:
            form = RutaForm(request.POST)
            if form.is_valid():
                ruta = form.save()
                logger.info(f'[RUTAS] Ruta creada: {ruta.codigo} - {ruta.nombre}')
                messages.success(request, f'Ruta "{ruta.codigo}" creada exitosamente.')
                return redirect('rutas')
            else:
                logger.warning(f'[RUTAS] Errores en formulario: {form.errors}')
                messages.error(request, 'Por favor corrige los errores en el formulario.')
        elif 'eliminar' in request.POST:
            ruta_id = request.POST.get('ruta_id')
            logger.info(f'[RUTAS] Eliminando ruta ID: {ruta_id}')
            
            try:
                ruta = get_object_or_404(Ruta, id=ruta_id)
                codigo = ruta.codigo
                
                # Verificar si hay usuarios asignados
                usuarios_asignados = UsuarioRuta.objects.filter(ruta=ruta).count()
                if usuarios_asignados > 0:
                    logger.warning(f'[RUTAS] No se puede eliminar: {usuarios_asignados} usuarios asignados')
                    messages.error(request, f'No se puede eliminar la ruta "{codigo}" porque tiene {usuarios_asignados} usuario(s) asignado(s).')
                else:
                    ruta.delete()
                    logger.info(f'[RUTAS] Ruta eliminada: {codigo}')
                    messages.success(request, f'Ruta "{codigo}" eliminada exitosamente.')
            except Exception as e:
                logger.error(f'[RUTAS] Error al eliminar: {e}')
                messages.error(request, f'Error al eliminar ruta: {str(e)}')
            
            return redirect('rutas')
        elif 'toggle_activa' in request.POST:
            ruta_id = request.POST.get('ruta_id')
            logger.info(f'[RUTAS] Toggle activa para ruta ID: {ruta_id}')
            
            ruta = get_object_or_404(Ruta, id=ruta_id)
            estado_anterior = ruta.activa
            ruta.activa = not ruta.activa
            ruta.save()
            
            estado = 'activada' if ruta.activa else 'desactivada'
            logger.info(f'[RUTAS] Ruta {ruta.codigo} {estado}')
            messages.success(request, f'Ruta "{ruta.codigo}" {estado} exitosamente.')
            return redirect('rutas')
    
    # Obtener estadísticas
    rutas_activas = rutas_list.filter(activa=True).count()
    rutas_inactivas = rutas_list.filter(activa=False).count()
    
    # Para cada ruta, obtener usuarios asignados y contar
    for ruta in rutas_list:
        relaciones = UsuarioRuta.objects.filter(ruta=ruta).select_related('usuario')
        ruta.usuarios_count = relaciones.count()
        ruta.usuarios_asignados = [rel.usuario for rel in relaciones]
    
    context = {
        'titulo': 'Gestión de Rutas',
        'rutas': rutas_list,
        'form': form,
        'rutas_activas': rutas_activas,
        'rutas_inactivas': rutas_inactivas,
    }
    return render(request, 'pagina/rutas.html', context)


@rol_requerido('ADMIN')
def parametros(request):
    """Gestión de Parámetros del Sistema"""
    from .models import Parametro
    from .forms import ParametroForm
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    
    parametros_list = Parametro.objects.all().order_by('clave')
    form = ParametroForm()
    
    if request.method == 'POST':
        logger.info(f'[PARAMETROS] ===== PETICIÓN POST RECIBIDA =====')
        logger.info(f'[PARAMETROS] Usuario: {request.user.username}')
        logger.info(f'[PARAMETROS] Datos POST: {dict(request.POST)}')
        
        if 'crear' in request.POST:
            form = ParametroForm(request.POST)
            if form.is_valid():
                parametro = form.save()
                logger.info(f'[PARAMETROS] Parámetro creado: {parametro.clave} = {parametro.valor}')
                messages.success(request, f'Parámetro "{parametro.clave}" creado exitosamente.')
                return redirect('parametros')
            else:
                logger.warning(f'[PARAMETROS] Errores en formulario: {form.errors}')
                messages.error(request, 'Por favor corrige los errores en el formulario.')
        elif 'eliminar' in request.POST:
            parametro_id = request.POST.get('parametro_id')
            logger.info(f'[PARAMETROS] Eliminando parámetro ID: {parametro_id}')
            
            try:
                parametro = get_object_or_404(Parametro, id=parametro_id)
                clave = parametro.clave
                parametro.delete()
                logger.info(f'[PARAMETROS] Parámetro eliminado: {clave}')
                messages.success(request, f'Parámetro "{clave}" eliminado exitosamente.')
            except Exception as e:
                logger.error(f'[PARAMETROS] Error al eliminar: {e}')
                messages.error(request, f'Error al eliminar parámetro: {str(e)}')
            
            return redirect('parametros')
    
    context = {
        'titulo': 'Parámetros del Sistema',
        'parametros': parametros_list,
        'form': form,
        'total_parametros': parametros_list.count(),
    }
    return render(request, 'pagina/parametros.html', context)


@rol_requerido('ADMIN')
def editar_parametro(request, parametro_id):
    """Editar un parámetro específico"""
    from .models import Parametro
    from .forms import ParametroForm
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    
    parametro = get_object_or_404(Parametro, id=parametro_id)
    
    if request.method == 'POST':
        form = ParametroForm(request.POST, instance=parametro)
        if form.is_valid():
            parametro_actualizado = form.save()
            logger.info(f'[PARAMETROS] Parámetro actualizado: {parametro_actualizado.clave}')
            messages.success(request, f'Parámetro "{parametro_actualizado.clave}" actualizado exitosamente.')
            return redirect('parametros')
        else:
            logger.warning(f'[PARAMETROS] Errores en formulario: {form.errors}')
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ParametroForm(instance=parametro)
    
    context = {
        'titulo': f'Editar Parámetro: {parametro.clave}',
        'parametro': parametro,
        'form': form,
    }
    return render(request, 'pagina/editar_parametro.html', context)


@rol_requerido('ADMIN')
def logs_sistema(request):
    """Logs del Sistema"""
    import os
    from pathlib import Path
    from django.http import JsonResponse
    
    BASE_DIR = Path(__file__).resolve().parent.parent.parent
    log_file = BASE_DIR / 'logs' / 'django.log'
    
    # Limpiar logs si se solicita
    if request.method == 'POST' and request.GET.get('limpiar') == '1':
        try:
            if log_file.exists():
                with open(log_file, 'w', encoding='utf-8') as f:
                    f.write('')
                logger.info('[LOGS_SISTEMA] Archivo de logs limpiado por usuario: {}'.format(request.user.username))
                return JsonResponse({'success': True, 'message': 'Logs limpiados exitosamente'})
            else:
                return JsonResponse({'success': False, 'message': 'El archivo de logs no existe'})
        except Exception as e:
            logger.error(f'[LOGS_SISTEMA] Error al limpiar logs: {e}')
            return JsonResponse({'success': False, 'message': f'Error al limpiar logs: {str(e)}'})
    
    logs = []
    total_lines = 0
    
    if log_file.exists():
        try:
            # Leer las últimas 500 líneas del archivo
            with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
                total_lines = len(lines)
                # Tomar las últimas 500 líneas
                logs = lines[-500:] if len(lines) > 500 else lines
                # Invertir para mostrar las más recientes primero
                logs.reverse()
        except Exception as e:
            logger.error(f'[LOGS_SISTEMA] Error leyendo archivo de logs: {e}')
            logs = [f'Error al leer el archivo de logs: {str(e)}']
    else:
        logs = ['El archivo de logs aún no existe. Los logs aparecerán aquí una vez que se generen.']
    
    # Filtrar logs si se proporciona un filtro
    filtro = request.GET.get('filtro', '').lower()
    if filtro:
        logs = [log for log in logs if filtro in log.lower()]
    
    context = {
        'titulo': 'Logs del Sistema',
        'logs': logs,
        'total_lines': total_lines,
        'showing_lines': len(logs),
        'filtro': request.GET.get('filtro', ''),
    }
    return render(request, 'pagina/logs_sistema.html', context)


@login_required
def perfil(request):
    """Perfil de Usuario"""
    from .models import UsuarioRuta
    from .forms import CambiarPasswordForm
    from django.contrib import messages
    from django.contrib.auth import update_session_auth_hash
    
    usuario = request.user
    
    # Obtener rutas asignadas al usuario
    rutas_asignadas = UsuarioRuta.objects.filter(usuario=usuario).select_related('ruta').order_by('ruta__codigo')
    
    # Formulario para cambiar contraseña
    password_form = CambiarPasswordForm(user=usuario)
    
    if request.method == 'POST' and 'cambiar_password' in request.POST:
        password_form = CambiarPasswordForm(user=usuario, data=request.POST)
        if password_form.is_valid():
            password_form.save()
            # Actualizar la sesión para que el usuario no se desloguee
            update_session_auth_hash(request, usuario)
            messages.success(request, 'Tu contraseña ha sido cambiada exitosamente.')
            logger.info(f'[PERFIL] Usuario {usuario.username} cambió su contraseña')
            return redirect('perfil')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    
    # Estadísticas del usuario
    total_rutas = rutas_asignadas.count()
    rutas_activas = sum(1 for rel in rutas_asignadas if rel.ruta.activa)
    
    context = {
        'titulo': 'Mi Perfil',
        'usuario': usuario,
        'rutas_asignadas': rutas_asignadas,
        'password_form': password_form,
        'total_rutas': total_rutas,
        'rutas_activas': rutas_activas,
    }
    return render(request, 'pagina/perfil.html', context)


@login_required
def notificaciones(request):
    """Configuración de Notificaciones"""
    context = {
        'titulo': 'Notificaciones',
        'contenido': 'Configurar alertas por email y preferencias de notificación.'
    }
    return render(request, 'pagina/pagina.html', context)
